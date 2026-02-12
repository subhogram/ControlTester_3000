"""
RCM Compliance Analyzer
Analyzes organization's Risk Control Matrix (RCM) against regulatory requirements.
Updated to handle multi-sheet Excel format with:
- Sheet 1: Company Controls
- Sheet 2: Company-CISM Controls

Format:
- Control Reference (e.g., A.5.1.1, C-CISM-4.1)
- Control Title
- Control Description
- Domain
- Sub Domain
"""

import os
import openpyxl
from typing import List, Dict, Any, Optional
from pathlib import Path
import logging
from datetime import datetime

logger = logging.getLogger(__name__)

# Lazy load LLM
_llm_cache = None

def get_llm(model: str):
    """Lazy load Ollama LLM."""
    global _llm_cache
    if _llm_cache is None or _llm_cache[0] != model:
        from langchain_community.llms import Ollama
        ollama_url = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
        _llm_cache = (model, Ollama(model=model, base_url=ollama_url, temperature=0.1))
    return _llm_cache[1]


# ============================================================================
# BACKWARD COMPATIBILITY FUNCTIONS
# (Keep these for existing code that imports from rcm_compliance_analyzer)
# ============================================================================

def load_document(file_path: str, filename: str = None) -> List[Any]:
    """
    Load document from file path (backward compatible).
    
    Args:
        file_path: Path to document file
        filename: Optional filename
    
    Returns:
        List of loaded documents with page_content attribute
    """
    from langchain_community.document_loaders import (
        PyPDFLoader, TextLoader, UnstructuredMarkdownLoader, CSVLoader
    )
    from langchain.schema import Document
    
    ext = Path(file_path).suffix.lower()
    
    try:
        if ext == '.pdf':
            loader = PyPDFLoader(file_path)
            return loader.load()
        elif ext == '.txt':
            loader = TextLoader(file_path)
            return loader.load()
        elif ext == '.md':
            loader = UnstructuredMarkdownLoader(file_path)
            return loader.load()
        elif ext == '.csv':
            loader = CSVLoader(file_path)
            return loader.load()
        elif ext in ['.xlsx', '.xls']:
            # For Excel files, parse as RCM and create Document objects
            logger.info(f"Loading Excel file as RCM: {file_path}")
            
            try:
                rcm_controls = parse_rcm_excel(file_path)
                documents = []
                
                # Create a Document object for each control
                for sheet_name, controls in rcm_controls.items():
                    for ctrl in controls:
                        # Build page_content from control information
                        page_content = f"""Control Reference: {ctrl.get('reference', 'N/A')}
Control Title: {ctrl.get('title', 'N/A')}
Control Description: {ctrl.get('description', 'N/A')}
Domain: {ctrl.get('domain', 'N/A')}
Sub Domain: {ctrl.get('subdomain', 'N/A')}
Sheet: {sheet_name}"""
                        
                        # Create Document with metadata
                        doc = Document(
                            page_content=page_content,
                            metadata={
                                'source': filename or file_path,
                                'sheet': sheet_name,
                                'control_id': ctrl.get('reference'),
                                'control_title': ctrl.get('title'),
                                'domain': ctrl.get('domain'),
                                'type': 'rcm_control'
                            }
                        )
                        documents.append(doc)
                
                logger.info(f"Created {len(documents)} Document objects from Excel RCM")
                return documents
                
            except Exception as e:
                logger.error(f"Failed to parse Excel as RCM: {e}")
                # Fallback: create a single document with error info
                return [Document(
                    page_content=f"Excel file: {filename or file_path}\nParsing failed: {str(e)}",
                    metadata={'source': filename or file_path, 'type': 'excel', 'error': str(e)}
                )]
        else:
            raise ValueError(f"Unsupported file type: {ext}")
        
    except Exception as e:
        logger.error(f"Failed to load document {file_path}: {e}")
        raise


def get_text_splitter():
    """Get text splitter for chunking documents (backward compatible)."""
    from langchain.text_splitter import RecursiveCharacterTextSplitter
    
    return RecursiveCharacterTextSplitter(
        chunk_size=1000,
        chunk_overlap=200,
        length_function=len
    )


class RegulatoryRequirementExtractor:
    """Extract regulatory requirements from documents (backward compatible)."""
    
    def __init__(self, model_name: str):
        self.model_name = model_name
        self.llm = get_llm(model_name)
    
    def extract(self, documents: List[Any]) -> List[Dict[str, Any]]:
        """Extract requirements from documents."""
        requirements = []
        
        for idx, doc in enumerate(documents[:10]):  # Limit to 10 docs
            try:
                content = doc.page_content if hasattr(doc, 'page_content') else str(doc)
                
                prompt = f"""Extract regulatory requirements from this document.
                
Document excerpt:
{content[:1000]}

List the key requirements, controls, or obligations.
Format: One requirement per line, starting with "REQ:"
"""
                
                response = self.llm.invoke(prompt)
                
                # Parse requirements
                for line in response.split('\n'):
                    if line.strip().startswith('REQ:'):
                        req_text = line.replace('REQ:', '').strip()
                        if req_text:
                            requirements.append({
                                'text': req_text,
                                'source_doc': idx,
                                'type': 'regulatory_requirement'
                            })
                
            except Exception as e:
                logger.error(f"Failed to extract from doc {idx}: {e}")
        
        return requirements
    
    def run(self, documents: List[Any]) -> List[Dict[str, Any]]:
        """Alias for extract() method (backward compatible)."""
        return self.extract(documents)


class RCMControlExtractor:
    """Extract controls from RCM documents (backward compatible)."""
    
    def __init__(self, model_name: str):
        self.model_name = model_name
        self.llm = get_llm(model_name)
    
    def extract(self, documents: List[Any]) -> List[Dict[str, Any]]:
        """Extract controls from RCM documents."""
        controls = []
        
        for doc in documents:
            try:
                # Check if document has metadata indicating it's from Excel
                if hasattr(doc, 'metadata') and doc.metadata.get('type') == 'rcm_control':
                    # Already parsed Excel control - extract from metadata and content
                    controls.append({
                        'control_id': doc.metadata.get('control_id', 'N/A'),
                        'title': doc.metadata.get('control_title', 'N/A'),
                        'description': doc.page_content.split('Control Description:')[1].split('\n')[0].strip() if 'Control Description:' in doc.page_content else '',
                        'domain': doc.metadata.get('domain', 'N/A'),
                        'subdomain': doc.metadata.get('subdomain', ''),
                        'sheet': doc.metadata.get('sheet', ''),
                        'type': 'rcm_control'
                    })
                else:
                    # Parse text-based RCM using LLM
                    content = doc.page_content if hasattr(doc, 'page_content') else str(doc)
                    
                    prompt = f"""Extract control information from this RCM document.

Document excerpt:
{content[:1000]}

List the controls with their IDs, titles, and descriptions.
Format: CONTROL_ID: [id] | TITLE: [title] | DESC: [description]
"""
                    
                    response = self.llm.invoke(prompt)
                    
                    # Parse controls (simple extraction)
                    for line in response.split('\n'):
                        if 'CONTROL_ID:' in line:
                            try:
                                parts = line.split('|')
                                ctrl_id = parts[0].split('CONTROL_ID:')[1].strip() if len(parts) > 0 else ''
                                title = parts[1].split('TITLE:')[1].strip() if len(parts) > 1 else ''
                                desc = parts[2].split('DESC:')[1].strip() if len(parts) > 2 else ''
                                
                                if ctrl_id:
                                    controls.append({
                                        'control_id': ctrl_id,
                                        'title': title,
                                        'description': desc,
                                        'type': 'rcm_control'
                                    })
                            except:
                                pass
                
            except Exception as e:
                logger.error(f"Failed to extract controls from document: {e}")
        
        logger.info(f"Extracted {len(controls)} total controls")
        return controls
    
    def run(self, documents: List[Any]) -> List[Dict[str, Any]]:
        """Alias for extract() method (backward compatible)."""
        return self.extract(documents)


class ComplianceAnalyzer:
    """Analyze compliance between RCM and regulatory requirements (backward compatible)."""
    
    def __init__(self, model_name: str):
        self.model_name = model_name
        self.llm = get_llm(model_name)
    
    def analyze(self, rcm_controls: List[Dict], reg_requirements: List[Dict]) -> List[Dict]:
        """Analyze compliance (simple list output)."""
        results = []
        
        for ctrl in rcm_controls[:50]:  # Limit for performance
            prompt = f"""Analyze if this control meets the regulatory requirements.

Control: {ctrl.get('control_id')} - {ctrl.get('title')}
Description: {ctrl.get('description', '')}

Requirements:
{chr(10).join([f"- {r.get('text', '')[:200]}" for r in reg_requirements[:5]])}

Status: COMPLIANT / PARTIAL / NON-COMPLIANT
Gaps: [list gaps]
"""
            
            try:
                response = self.llm.invoke(prompt)
                
                status = 'UNKNOWN'
                if 'COMPLIANT' in response:
                    status = 'COMPLIANT'
                elif 'PARTIAL' in response:
                    status = 'PARTIAL'
                elif 'NON-COMPLIANT' in response:
                    status = 'NON-COMPLIANT'
                
                results.append({
                    'control_id': ctrl.get('control_id'),
                    'control_title': ctrl.get('title'),
                    'status': status,
                    'analysis': response[:300]
                })
                
            except Exception as e:
                logger.error(f"Analysis failed for {ctrl.get('control_id')}: {e}")
        
        return results
    
    def analyze_compliance(self, requirements: List[Dict], controls: List[Dict]) -> Dict[str, Any]:
        """
        Comprehensive compliance analysis with structured output.
        
        Args:
            requirements: List of regulatory requirements
            controls: List of RCM controls
        
        Returns:
            {
                'compliance_stats': {...},
                'gaps': [...],
                'domain_analyses': {...},
                'detailed_results': [...],
                'risk_level': 'HIGH/MEDIUM/LOW'
            }
        """
        logger.info(f"Analyzing compliance for {len(controls)} controls against {len(requirements)} requirements")
        
        detailed_results = []
        gaps = []
        domain_data = {}
        
        # Analyze each control
        for ctrl in controls[:50]:  # Limit for performance
            control_id = ctrl.get('control_id', 'N/A')
            control_title = ctrl.get('title', 'N/A')
            control_desc = ctrl.get('description', '')
            domain = ctrl.get('domain', 'Unknown')
            
            # Build analysis prompt
            req_text = '\n'.join([f"- {r.get('text', '')[:150]}" for r in requirements[:5]])
            
            prompt = f"""Analyze compliance for this control against regulatory requirements.

Control ID: {control_id}
Control: {control_title}
Description: {control_desc}

Regulatory Requirements:
{req_text}

Provide:
1. Status: COMPLIANT / PARTIAL / NON-COMPLIANT / UNABLE_TO_ASSESS
2. Compliance Score: 0-100
3. Gaps: What's missing or inadequate
4. Impact: HIGH / MEDIUM / LOW

Keep response structured and concise.
"""
            
            try:
                response = self.llm.invoke(prompt)
                
                # Parse status
                status = 'UNABLE_TO_ASSESS'
                if 'COMPLIANT' in response and 'NON-COMPLIANT' not in response:
                    status = 'COMPLIANT'
                elif 'PARTIAL' in response:
                    status = 'PARTIAL'
                elif 'NON-COMPLIANT' in response:
                    status = 'NON-COMPLIANT'
                
                # Parse score (look for numbers)
                score = 50
                import re
                score_match = re.search(r'Score:?\s*(\d+)', response, re.IGNORECASE)
                if score_match:
                    score = min(100, max(0, int(score_match.group(1))))
                elif status == 'COMPLIANT':
                    score = 90
                elif status == 'PARTIAL':
                    score = 60
                elif status == 'NON-COMPLIANT':
                    score = 30
                
                # Parse impact
                impact = 'MEDIUM'
                if 'HIGH' in response.upper():
                    impact = 'HIGH'
                elif 'LOW' in response.upper():
                    impact = 'LOW'
                
                # Extract gaps
                gap_text = ''
                if 'Gaps:' in response:
                    gap_text = response.split('Gaps:')[1].split('\n')[0].strip()
                elif status != 'COMPLIANT':
                    gap_text = response[:200]
                
                # Store detailed result
                result = {
                    'control_id': control_id,
                    'control_title': control_title,
                    'control_description': control_desc[:150],
                    'domain': domain,
                    'status': status,
                    'compliance_score': score,
                    'analysis': response[:400],
                    'impact': impact
                }
                detailed_results.append(result)
                
                # Track gaps
                if status in ['PARTIAL', 'NON-COMPLIANT']:
                    gaps.append({
                        'control_id': control_id,
                        'control': control_title,
                        'domain': domain,
                        'gap': gap_text or f"{status}: {response[:150]}",
                        'impact': impact,
                        'status': status
                    })
                
                # Organize by domain
                if domain not in domain_data:
                    domain_data[domain] = {
                        'controls': [],
                        'total_controls': 0,
                        'compliant': 0,
                        'partial': 0,
                        'non_compliant': 0,
                        'avg_score': 0,
                        'findings': []
                    }
                
                domain_data[domain]['controls'].append(result)
                domain_data[domain]['total_controls'] += 1
                
                if status == 'COMPLIANT':
                    domain_data[domain]['compliant'] += 1
                elif status == 'PARTIAL':
                    domain_data[domain]['partial'] += 1
                elif status == 'NON-COMPLIANT':
                    domain_data[domain]['non_compliant'] += 1
                
            except Exception as e:
                logger.error(f"Analysis failed for {control_id}: {e}")
                detailed_results.append({
                    'control_id': control_id,
                    'control_title': control_title,
                    'status': 'ERROR',
                    'analysis': f"Analysis error: {str(e)[:100]}",
                    'compliance_score': 0
                })
        
        # Calculate overall stats
        total = len(detailed_results)
        compliant = sum(1 for r in detailed_results if r.get('status') == 'COMPLIANT')
        partial = sum(1 for r in detailed_results if r.get('status') == 'PARTIAL')
        non_compliant = sum(1 for r in detailed_results if r.get('status') == 'NON-COMPLIANT')
        unable = sum(1 for r in detailed_results if r.get('status') == 'UNABLE_TO_ASSESS')
        errors = sum(1 for r in detailed_results if r.get('status') == 'ERROR')
        
        # Average compliance score
        scores = [r.get('compliance_score', 0) for r in detailed_results if r.get('compliance_score')]
        avg_score = sum(scores) / len(scores) if scores else 0
        
        # Calculate domain averages
        for domain, data in domain_data.items():
            domain_scores = [c.get('compliance_score', 0) for c in data['controls']]
            data['avg_score'] = sum(domain_scores) / len(domain_scores) if domain_scores else 0
            
            # Add findings
            data['findings'] = [
                {
                    'control_id': c['control_id'],
                    'status': c['status'],
                    'summary': c['analysis'][:150]
                }
                for c in data['controls'][:5]  # Top 5 findings per domain
            ]
        
        # Determine overall risk level
        if avg_score >= 80:
            risk_level = 'LOW'
        elif avg_score >= 60:
            risk_level = 'MEDIUM'
        else:
            risk_level = 'HIGH'
        
        return {
            'compliance_stats': {
                'total': total,
                'compliant': compliant,
                'partial': partial,
                'non_compliant': non_compliant,
                'unable_to_assess': unable,
                'errors': errors,
                'compliance_rate': (compliant / total * 100) if total > 0 else 0,
                'avg_compliance_score': round(avg_score, 2)
            },
            'gaps': gaps,
            'domain_analyses': domain_data,
            'detailed_results': detailed_results,
            'risk_level': risk_level
        }


class RemediationSuggester:
    """Suggest remediations for gaps (backward compatible)."""
    
    def __init__(self, model_name: str):
        self.model_name = model_name
        self.llm = get_llm(model_name)
    
    def suggest(self, compliance_results: List[Dict]) -> List[Dict]:
        """Suggest remediations."""
        suggestions = []
        
        non_compliant = [r for r in compliance_results if r.get('status') == 'NON-COMPLIANT']
        
        for result in non_compliant[:10]:
            prompt = f"""Suggest remediation for this gap:

Control: {result.get('control_id')}
Issue: {result.get('analysis', '')}

Provide specific, actionable steps to achieve compliance.
"""
            
            try:
                response = self.llm.invoke(prompt)
                
                suggestions.append({
                    'control_id': result.get('control_id'),
                    'recommendation': response[:300]
                })
                
            except Exception as e:
                logger.error(f"Remediation suggestion failed: {e}")
        
        return suggestions
    
    def generate_suggestions(self, gaps: List[Dict], domain_analyses: Dict[str, Any]) -> Dict[str, Any]:
        """Generate comprehensive remediation suggestions (backward compatible)."""
        
        all_suggestions = []
        priority_suggestions = []
        domain_suggestions = {}
        
        # Process gaps
        for gap in gaps[:20]:  # Limit to 20 for performance
            prompt = f"""Provide remediation for this compliance gap:

Control ID: {gap.get('control_id', 'N/A')}
Control: {gap.get('control', 'N/A')}
Gap: {gap.get('gap', 'N/A')}
Impact: {gap.get('impact', 'MEDIUM')}

Provide:
1. Specific remediation steps
2. Priority (HIGH/MEDIUM/LOW)
3. Estimated effort
4. Resources needed

Keep response concise.
"""
            
            try:
                response = self.llm.invoke(prompt)
                
                # Parse priority
                priority = 'MEDIUM'
                if 'HIGH' in response.upper():
                    priority = 'HIGH'
                elif 'LOW' in response.upper():
                    priority = 'LOW'
                
                suggestion = {
                    'control_id': gap.get('control_id'),
                    'control': gap.get('control', ''),
                    'gap': gap.get('gap', ''),
                    'remediation': response[:500],
                    'priority': priority,
                    'domain': gap.get('domain', 'Unknown')
                }
                
                all_suggestions.append(suggestion)
                
                if priority == 'HIGH':
                    priority_suggestions.append(suggestion)
                
                # Group by domain
                domain = gap.get('domain', 'Unknown')
                if domain not in domain_suggestions:
                    domain_suggestions[domain] = []
                domain_suggestions[domain].append(suggestion)
                
            except Exception as e:
                logger.error(f"Failed to generate suggestion: {e}")
        
        return {
            'all_suggestions': all_suggestions,
            'priority_suggestions': priority_suggestions,
            'domain_suggestions': domain_suggestions,
            'total_suggestions': len(all_suggestions),
            'high_priority_count': len(priority_suggestions)
        }
    
    def generate_suggestions(self, gaps: List[Any], domain_analyses: Dict[str, Any]) -> Dict[str, List[Dict]]:
        """
        Generate remediation suggestions organized by domain.
        
        Args:
            gaps: List of identified gaps
            domain_analyses: Domain-specific analysis results
        
        Returns:
            Dict mapping domain names to lists of suggestions
        """
        suggestions_by_domain = {}
        
        try:
            # Process domain analyses
            for domain, analysis in domain_analyses.items():
                domain_suggestions = []
                
                # Get gaps for this domain
                domain_gaps = analysis.get('gaps', []) if isinstance(analysis, dict) else []
                
                for gap in domain_gaps[:5]:  # Limit to 5 per domain
                    try:
                        gap_text = gap if isinstance(gap, str) else gap.get('description', str(gap))
                        
                        prompt = f"""Provide remediation for this compliance gap:

Domain: {domain}
Gap: {gap_text}

Provide:
1. Root cause
2. Specific remediation steps
3. Priority (HIGH/MEDIUM/LOW)
4. Estimated effort

Keep response concise (max 200 words).
"""
                        
                        response = self.llm.invoke(prompt)
                        
                        # Parse priority
                        priority = 'MEDIUM'
                        if 'HIGH' in response.upper():
                            priority = 'HIGH'
                        elif 'LOW' in response.upper():
                            priority = 'LOW'
                        
                        domain_suggestions.append({
                            'gap': gap_text[:200],
                            'recommendation': response[:400],
                            'priority': priority,
                            'domain': domain
                        })
                        
                    except Exception as e:
                        logger.error(f"Failed to generate suggestion for gap in {domain}: {e}")
                
                if domain_suggestions:
                    suggestions_by_domain[domain] = domain_suggestions
            
            # Also process general gaps not in domain analyses
            if gaps and not domain_analyses:
                general_suggestions = []
                
                for gap in gaps[:10]:
                    try:
                        gap_text = gap if isinstance(gap, str) else str(gap)
                        
                        prompt = f"""Provide remediation for this gap: {gap_text}

Provide specific, actionable steps.
"""
                        response = self.llm.invoke(prompt)
                        
                        general_suggestions.append({
                            'gap': gap_text[:200],
                            'recommendation': response[:300]
                        })
                        
                    except Exception as e:
                        logger.error(f"Failed to generate suggestion: {e}")
                
                if general_suggestions:
                    suggestions_by_domain['General'] = general_suggestions
            
        except Exception as e:
            logger.error(f"Suggestion generation failed: {e}")
        
        return suggestions_by_domain


class ComplianceReportGenerator:
    """Generate compliance reports (backward compatible)."""
    
    def __init__(self, model_name: str):
        self.model_name = model_name
    
    def generate(self, 
                 rcm_controls: List[Dict],
                 compliance_results: List[Dict],
                 remediations: List[Dict]) -> str:
        """Generate report."""
        
        report = []
        report.append("=" * 80)
        report.append("RCM COMPLIANCE REPORT")
        report.append("=" * 80)
        report.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Summary
        total = len(compliance_results)
        compliant = sum(1 for r in compliance_results if r.get('status') == 'COMPLIANT')
        partial = sum(1 for r in compliance_results if r.get('status') == 'PARTIAL')
        non_compliant = sum(1 for r in compliance_results if r.get('status') == 'NON-COMPLIANT')
        
        report.append(f"\n\nTotal Controls: {total}")
        report.append(f"Compliant: {compliant}")
        report.append(f"Partially Compliant: {partial}")
        report.append(f"Non-Compliant: {non_compliant}")
        
        # Details
        report.append("\n\n" + "=" * 80)
        report.append("DETAILED RESULTS")
        report.append("=" * 80)
        
        for result in compliance_results:
            report.append(f"\n{result.get('control_id')}: {result.get('status')}")
            report.append(f"  {result.get('analysis', '')[:200]}")
        
        # Remediations
        if remediations:
            report.append("\n\n" + "=" * 80)
            report.append("REMEDIATION RECOMMENDATIONS")
            report.append("=" * 80)
            
            for rem in remediations:
                report.append(f"\n{rem.get('control_id')}:")
                report.append(f"  {rem.get('recommendation', '')[:200]}")
        
        report.append("\n\n" + "=" * 80)
        report.append("END OF REPORT")
        report.append("=" * 80)
        
        return "\n".join(report)
    
    def generate_executive_summary(self, analysis: Dict[str, Any], suggestions: Dict[str, Any]) -> str:
        """Generate executive summary (backward compatible)."""
        
        summary = []
        summary.append("=" * 80)
        summary.append("EXECUTIVE SUMMARY")
        summary.append("=" * 80)
        summary.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # Overall metrics
        summary.append("OVERALL COMPLIANCE STATUS")
        summary.append("-" * 80)
        
        compliance_stats = analysis.get('compliance_stats', {})
        total = compliance_stats.get('total', 0)
        compliant = compliance_stats.get('compliant', 0)
        partial = compliance_stats.get('partial', 0)
        non_compliant = compliance_stats.get('non_compliant', 0)
        
        if total > 0:
            compliance_rate = (compliant / total) * 100
            summary.append(f"Total Controls Analyzed: {total}")
            summary.append(f"Compliant: {compliant} ({compliant/total*100:.1f}%)")
            summary.append(f"Partially Compliant: {partial} ({partial/total*100:.1f}%)")
            summary.append(f"Non-Compliant: {non_compliant} ({non_compliant/total*100:.1f}%)")
            summary.append(f"\nOverall Compliance Rate: {compliance_rate:.1f}%")
        
        # Risk assessment
        summary.append("\n\nRISK ASSESSMENT")
        summary.append("-" * 80)
        
        risk_level = analysis.get('risk_level', 'MEDIUM')
        summary.append(f"Overall Risk Level: {risk_level}")
        
        gaps = analysis.get('gaps', [])
        if gaps:
            high_risk_gaps = [g for g in gaps if g.get('impact') == 'HIGH']
            summary.append(f"Critical Gaps: {len(high_risk_gaps)}")
            summary.append(f"Total Gaps: {len(gaps)}")
        
        # Remediation overview
        summary.append("\n\nREMEDIATION OVERVIEW")
        summary.append("-" * 80)
        
        total_suggestions = suggestions.get('total_suggestions', 0)
        high_priority = suggestions.get('high_priority_count', 0)
        
        summary.append(f"Total Remediation Actions: {total_suggestions}")
        summary.append(f"High Priority Actions: {high_priority}")
        
        # Priority actions
        if suggestions.get('priority_suggestions'):
            summary.append("\n\nTOP PRIORITY ACTIONS")
            summary.append("-" * 80)
            
            for idx, sug in enumerate(suggestions['priority_suggestions'][:5], 1):
                summary.append(f"\n{idx}. {sug.get('control_id')} - {sug.get('control', '')[:60]}")
                summary.append(f"   Gap: {sug.get('gap', '')[:100]}")
                summary.append(f"   Action: {sug.get('remediation', '')[:150]}")
        
        summary.append("\n\n" + "=" * 80)
        
        return "\n".join(summary)
    
    def generate_domain_report(self, domain: str, domain_analysis: Dict[str, Any], domain_suggestions: List[Dict]) -> str:
        """Generate domain-specific report (backward compatible)."""
        
        report = []
        report.append("=" * 80)
        report.append(f"DOMAIN REPORT: {domain}")
        report.append("=" * 80)
        report.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # Domain statistics
        report.append("DOMAIN STATISTICS")
        report.append("-" * 80)
        
        total_controls = domain_analysis.get('total_controls', 0)
        compliant = domain_analysis.get('compliant', 0)
        non_compliant = domain_analysis.get('non_compliant', 0)
        
        report.append(f"Total Controls: {total_controls}")
        if total_controls > 0:
            report.append(f"Compliant: {compliant} ({compliant/total_controls*100:.1f}%)")
            report.append(f"Non-Compliant: {non_compliant} ({non_compliant/total_controls*100:.1f}%)")
        
        # Key findings
        report.append("\n\nKEY FINDINGS")
        report.append("-" * 80)
        
        findings = domain_analysis.get('findings', [])
        if findings:
            for idx, finding in enumerate(findings[:5], 1):
                report.append(f"\n{idx}. {finding.get('control_id')}: {finding.get('status')}")
                report.append(f"   {finding.get('summary', '')[:200]}")
        else:
            report.append("No specific findings for this domain.")
        
        # Remediation actions
        if domain_suggestions:
            report.append("\n\nREMEDIATION ACTIONS")
            report.append("-" * 80)
            
            for idx, sug in enumerate(domain_suggestions[:10], 1):
                report.append(f"\n{idx}. {sug.get('control_id')}")
                report.append(f"   Gap: {sug.get('gap', '')[:150]}")
                report.append(f"   Remediation: {sug.get('remediation', '')[:200]}")
                report.append(f"   Priority: {sug.get('priority', 'MEDIUM')}")
        
        report.append("\n\n" + "=" * 80)
        
        return "\n".join(report)
    
    def generate_executive_summary(self, analysis: Dict[str, Any], suggestions: Dict[str, List[Dict]]) -> str:
        """
        Generate executive summary.
        
        Args:
            analysis: Compliance analysis results
            suggestions: Remediation suggestions by domain
        
        Returns:
            Executive summary text
        """
        summary = []
        summary.append("=" * 80)
        summary.append("EXECUTIVE SUMMARY - RCM COMPLIANCE ANALYSIS")
        summary.append("=" * 80)
        summary.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Overall metrics
        summary.append("\n\nOVERALL COMPLIANCE STATUS")
        summary.append("-" * 80)
        
        overall_score = analysis.get('overall_compliance_score', 0)
        total_controls = analysis.get('total_controls', 0)
        compliant_count = analysis.get('compliant_controls', 0)
        
        summary.append(f"Total Controls Analyzed: {total_controls}")
        summary.append(f"Overall Compliance Score: {overall_score:.1f}%")
        summary.append(f"Compliant Controls: {compliant_count}")
        
        # Risk assessment
        summary.append("\n\nRISK ASSESSMENT")
        summary.append("-" * 80)
        
        risk_level = analysis.get('risk_level', 'UNKNOWN')
        summary.append(f"Risk Level: {risk_level}")
        
        # Key findings
        summary.append("\n\nKEY FINDINGS")
        summary.append("-" * 80)
        
        gaps = analysis.get('gaps', [])
        if gaps:
            summary.append(f"\nTotal Gaps Identified: {len(gaps)}")
            summary.append("\nTop Critical Gaps:")
            for idx, gap in enumerate(gaps[:5], 1):
                gap_text = gap if isinstance(gap, str) else gap.get('description', str(gap))
                summary.append(f"  {idx}. {gap_text[:150]}")
        else:
            summary.append("\nNo critical gaps identified.")
        
        # Remediation overview
        summary.append("\n\nREMEDIATION RECOMMENDATIONS")
        summary.append("-" * 80)
        
        total_suggestions = sum(len(suggs) for suggs in suggestions.values())
        summary.append(f"\nTotal Recommendations: {total_suggestions}")
        
        if suggestions:
            summary.append("\nBy Domain:")
            for domain, suggs in suggestions.items():
                summary.append(f"  - {domain}: {len(suggs)} recommendations")
        
        # Domain summary
        domain_analyses = analysis.get('domain_analyses', {})
        if domain_analyses:
            summary.append("\n\nDOMAIN COMPLIANCE SUMMARY")
            summary.append("-" * 80)
            
            for domain, domain_data in domain_analyses.items():
                if isinstance(domain_data, dict):
                    domain_score = domain_data.get('compliance_score', 0)
                    summary.append(f"\n{domain}: {domain_score:.1f}%")
        
        summary.append("\n\n" + "=" * 80)
        summary.append("END OF EXECUTIVE SUMMARY")
        summary.append("=" * 80)
        
        return "\n".join(summary)
    
    def generate_domain_report(self, domain: str, domain_analysis: Dict[str, Any], domain_suggestions: List[Dict]) -> str:
        """
        Generate domain-specific report.
        
        Args:
            domain: Domain name
            domain_analysis: Analysis results for this domain
            domain_suggestions: Remediation suggestions for this domain
        
        Returns:
            Domain report text
        """
        report = []
        report.append("=" * 80)
        report.append(f"DOMAIN REPORT: {domain.upper()}")
        report.append("=" * 80)
        report.append(f"\nGenerated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Domain metrics
        report.append("\n\nDOMAIN METRICS")
        report.append("-" * 80)
        
        if isinstance(domain_analysis, dict):
            total_controls = domain_analysis.get('total_controls', 0)
            compliant = domain_analysis.get('compliant_controls', 0)
            compliance_score = domain_analysis.get('compliance_score', 0)
            
            report.append(f"Total Controls: {total_controls}")
            report.append(f"Compliant Controls: {compliant}")
            report.append(f"Compliance Score: {compliance_score:.1f}%")
            
            # Gaps
            gaps = domain_analysis.get('gaps', [])
            if gaps:
                report.append(f"\n\nIDENTIFIED GAPS ({len(gaps)})")
                report.append("-" * 80)
                
                for idx, gap in enumerate(gaps[:10], 1):
                    gap_text = gap if isinstance(gap, str) else gap.get('description', str(gap))
                    report.append(f"\n{idx}. {gap_text}")
        
        # Recommendations
        if domain_suggestions:
            report.append(f"\n\nRECOMMENDATIONS ({len(domain_suggestions)})")
            report.append("-" * 80)
            
            for idx, suggestion in enumerate(domain_suggestions, 1):
                report.append(f"\n{idx}. Gap: {suggestion.get('gap', 'N/A')[:100]}")
                report.append(f"   Priority: {suggestion.get('priority', 'MEDIUM')}")
                report.append(f"   Recommendation: {suggestion.get('recommendation', 'N/A')[:200]}")
        
        report.append("\n\n" + "=" * 80)
        report.append(f"END OF {domain.upper()} REPORT")
        report.append("=" * 80)
        
        return "\n".join(report)


# ============================================================================
# MAIN RCM COMPLIANCE FUNCTIONS (Updated for Excel)
# ============================================================================


def parse_rcm_excel(file_path: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Parse RCM Excel file with multiple sheets.
    
    Expected format:
    - Row 1: Headers (Control Reference, Control Title, Control Description, Domain, Sub Domain)
    - Row 2+: Control data
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Dict with sheet names as keys and list of controls as values
    """
    logger.info(f"Parsing RCM Excel file: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_controls = {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logger.info(f"Processing sheet: {sheet_name}")
            
            # Find header row (usually row 1)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(None)
            
            logger.info(f"Headers found: {headers}")
            
            # Map expected columns (flexible matching)
            col_map = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                    
                header_lower = header.lower()
                
                if 'control reference' in header_lower or 'control ref' in header_lower:
                    col_map['reference'] = idx
                elif 'control title' in header_lower or 'title' in header_lower:
                    col_map['title'] = idx
                elif 'control description' in header_lower or 'description' in header_lower:
                    col_map['description'] = idx
                elif header_lower == 'domain':
                    col_map['domain'] = idx
                elif 'sub domain' in header_lower or 'subdomain' in header_lower or 'sub-domain' in header_lower:
                    col_map['subdomain'] = idx
            
            logger.info(f"Column mapping: {col_map}")
            
            # Validate required columns
            if 'reference' not in col_map:
                logger.warning(f"Sheet {sheet_name}: Missing 'Control Reference' column, skipping")
                continue
            
            # Extract controls (skip header row)
            controls = []
            for row_idx in range(2, ws.max_row + 1):
                row = ws[row_idx]
                
                # Build control dict
                control = {
                    'sheet': sheet_name,
                    'row': row_idx
                }
                
                # Extract mapped columns
                for field, col_idx in col_map.items():
                    if col_idx < len(row):
                        value = row[col_idx].value
                        if value:
                            # Clean up text (remove extra whitespace, newlines)
                            control[field] = str(value).strip().replace('\n', ' ')
                
                # Only include rows with at least a reference
                if 'reference' in control and control['reference']:
                    controls.append(control)
            
            all_controls[sheet_name] = controls
            logger.info(f"Extracted {len(controls)} controls from {sheet_name}")
        
        # Log summary
        total = sum(len(controls) for controls in all_controls.values())
        logger.info(f"Total controls extracted: {total} across {len(all_controls)} sheets")
        
        return all_controls
        
    except Exception as e:
        logger.error(f"Failed to parse RCM Excel: {e}")
        raise ValueError(f"RCM Excel parsing failed: {e}")


def analyze_rcm_compliance(
    rcm_file_path: str,
    regulatory_docs: List[str],
    model_name: str,
    regulatory_vectorstores: Optional[List[Any]] = None
) -> Dict[str, Any]:

    logger.info("Starting Enhanced RCM compliance analysis")

    try:
        # -----------------------------
        # STEP 1 – Parse RCM
        # -----------------------------
        rcm_controls_by_sheet = parse_rcm_excel(rcm_file_path)

        all_controls = []
        for sheet_name, controls in rcm_controls_by_sheet.items():
            for ctrl in controls:
                ctrl["sheet"] = sheet_name
                all_controls.append(ctrl)

        if not all_controls:
            raise ValueError("No controls found in RCM file.")

        # -----------------------------
        # STEP 2 – Extract Regulatory Obligations
        # -----------------------------
        llm = get_llm(model_name)

        regulatory_obligations = []

        from langchain_community.document_loaders import (
            PyPDFLoader, TextLoader, UnstructuredMarkdownLoader
        )

        for reg_doc_path in regulatory_docs:

            ext = Path(reg_doc_path).suffix.lower()

            if ext == ".pdf":
                loader = PyPDFLoader(reg_doc_path)
            elif ext in [".txt", ".md"]:
                loader = TextLoader(reg_doc_path)
            else:
                continue

            docs = loader.load()

            full_text = "\n".join([d.page_content for d in docs])[:12000]

            prompt = f"""
You are a regulatory compliance expert.

Extract ONLY enforceable regulatory requirements or obligations.

Ignore background text.

Return STRICT JSON list format:
[
  {{
    "requirement": "...",
    "domain": "...",
    "criticality": "HIGH/MEDIUM/LOW"
  }}
]

Text:
{full_text}
"""

            response = llm.invoke(prompt)

            try:
                import json
                extracted = json.loads(response)
                for item in extracted:
                    item["source"] = Path(reg_doc_path).name
                    regulatory_obligations.append(item)
            except Exception:
                logger.warning("Failed to parse structured obligations from LLM output.")

        if not regulatory_obligations:
            raise ValueError("No regulatory obligations extracted.")

        # -----------------------------
        # STEP 3 – Group By Domain
        # -----------------------------
        from collections import defaultdict

        obligations_by_domain = defaultdict(list)
        controls_by_domain = defaultdict(list)

        for req in regulatory_obligations:
            domain = req.get("domain", "General")
            obligations_by_domain[domain].append(req)

        for ctrl in all_controls:
            domain = ctrl.get("domain", "General")
            controls_by_domain[domain].append(ctrl)

        # -----------------------------
        # STEP 4 – Domain Level Analysis
        # -----------------------------
        compliance_results = []

        domain_scores = {}

        for domain, requirements in obligations_by_domain.items():

            controls = controls_by_domain.get(domain, [])

            req_text = "\n".join(
                [f"- {r['requirement']}" for r in requirements]
            )

            ctrl_text = "\n".join(
                [f"- {c.get('reference')} : {c.get('description','')}"
                 for c in controls]
            )

            prompt = f"""
You are a senior IT auditor.

Domain: {domain}

Regulatory Requirements:
{req_text}

RCM Controls:
{ctrl_text}

Tasks:
1. Map requirements to controls.
2. Identify missing requirements.
3. Identify weak or partial controls.
4. Calculate compliance score (0-100).
5. Provide remediation recommendations.

Return STRICT JSON:
{{
  "score": 0-100,
  "missing_requirements": [],
  "weak_controls": [],
  "recommendations": []
}}
"""

            response = llm.invoke(prompt)

            try:
                import json
                domain_analysis = json.loads(response)
            except Exception:
                domain_analysis = {
                    "score": 50,
                    "missing_requirements": [],
                    "weak_controls": [],
                    "recommendations": []
                }

            domain_scores[domain] = domain_analysis.get("score", 50)

            # Assign per control status for compatibility
            for ctrl in controls:
                status = "COMPLIANT"

                if ctrl.get("reference") in domain_analysis.get("weak_controls", []):
                    status = "PARTIAL"

                compliance_results.append({
                    "control_reference": ctrl.get("reference"),
                    "control_title": ctrl.get("title"),
                    "control_description": ctrl.get("description", "")[:200],
                    "domain": domain,
                    "subdomain": ctrl.get("subdomain"),
                    "sheet": ctrl.get("sheet"),
                    "compliance_status": status,
                    "gaps": ", ".join(domain_analysis.get("missing_requirements", []))[:300],
                    "recommendation": ", ".join(domain_analysis.get("recommendations", []))[:300],
                    "regulatory_matches": len(requirements)
                })

        # -----------------------------
        # STEP 5 – Summary Metrics
        # -----------------------------
        total = len(compliance_results)
        compliant = sum(1 for r in compliance_results if r["compliance_status"] == "COMPLIANT")
        partial = sum(1 for r in compliance_results if r["compliance_status"] == "PARTIAL")
        non_compliant = total - compliant - partial

        overall_score = sum(domain_scores.values()) / len(domain_scores)

        summary = {
            "total_controls": len(all_controls),
            "controls_analyzed": total,
            "compliant": compliant,
            "partial_compliant": partial,
            "non_compliant": non_compliant,
            "unable_to_assess": 0,
            "errors": 0,
            "overall_compliance_score": round(overall_score, 2),
            "risk_level": "HIGH" if overall_score < 50 else "MEDIUM" if overall_score < 80 else "LOW"
        }

        # -----------------------------
        # STEP 6 – Generate Comprehensive Report
        # -----------------------------
        final_report = generate_compliance_report(
            rcm_controls=rcm_controls_by_sheet,
            compliance_results=compliance_results,
            summary=summary,
            regulatory_docs=[Path(p).name for p in regulatory_docs]
        )

        return {
            "success": True,
            "rcm_structure": {s: len(c) for s, c in rcm_controls_by_sheet.items()},
            "compliance_analysis": {
                "overall_metrics": summary,
                "risk_score": round(100 - overall_score, 2),
                "detailed_results": compliance_results
            },
            "final_report": final_report,
            "timestamp": datetime.now().isoformat()
        }

    except Exception as e:
        logger.error(f"Enhanced RCM compliance analysis failed: {e}")
        return {
            "success": False,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

def generate_compliance_report(
    rcm_controls: Dict[str, List[Dict]],
    compliance_results: List[Dict],
    summary: Dict,
    regulatory_docs: List[str]
) -> str:

    from collections import defaultdict
    from datetime import datetime

    report = []

    report.append("=" * 110)
    report.append("ENTERPRISE TECHNOLOGY RISK & CONTROL COMPLIANCE REPORT")
    report.append("=" * 110)

    report.append(f"\nGenerated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append(f"Regulatory Frameworks Reviewed: {', '.join(regulatory_docs)}")

    # ------------------------------------------------------------------
    # EXECUTIVE SUMMARY
    # ------------------------------------------------------------------
    report.append("\n" + "=" * 110)
    report.append("EXECUTIVE SUMMARY")
    report.append("=" * 110)

    report.append(f"\nTotal Controls in RCM           : {summary['total_controls']}")
    report.append(f"Controls Analyzed              : {summary['controls_analyzed']}")
    report.append(f"Overall Compliance Score       : {summary['overall_compliance_score']}%")
    report.append(f"Overall Risk Rating            : {summary['risk_level']}")

    report.append("\nCompliance Distribution:")
    report.append(f"  ✔ Fully Compliant            : {summary['compliant']}")
    report.append(f"  ⚠ Partially Compliant        : {summary['partial_compliant']}")
    report.append(f"  ✖ Non-Compliant              : {summary['non_compliant']}")

    # ------------------------------------------------------------------
    # GROUP RESULTS BY DOMAIN
    # ------------------------------------------------------------------
    domain_map = defaultdict(list)
    for result in compliance_results:
        domain_map[result.get("domain", "General")].append(result)

    # ------------------------------------------------------------------
    # DOMAIN-WISE COMPREHENSIVE ANALYSIS
    # ------------------------------------------------------------------
    report.append("\n" + "=" * 110)
    report.append("DOMAIN-WISE COMPREHENSIVE COMPLIANCE ASSESSMENT")
    report.append("=" * 110)

    for domain, controls in sorted(domain_map.items()):

        report.append(f"\n\nDOMAIN: {domain.upper()}")
        report.append("-" * 110)

        total = len(controls)
        compliant = sum(1 for c in controls if c["compliance_status"] == "COMPLIANT")
        partial = sum(1 for c in controls if c["compliance_status"] == "PARTIAL")
        non_compliant = sum(1 for c in controls if c["compliance_status"] == "NON-COMPLIANT")

        domain_score = round(
            ((compliant + 0.5 * partial) / total) * 100, 2
        ) if total else 0

        report.append(f"Domain Compliance Score : {domain_score}%")
        report.append(f"Total Controls          : {total}")
        report.append(f"Compliant               : {compliant}")
        report.append(f"Partially Compliant     : {partial}")
        report.append(f"Non-Compliant           : {non_compliant}")

        report.append("\nCONTROL-BY-CONTROL ASSESSMENT:")
        report.append("-" * 110)

        for idx, ctrl in enumerate(controls, 1):

            report.append(f"\n{idx}. CONTROL REFERENCE : {ctrl.get('control_reference', 'N/A')}")
            report.append(f"   CONTROL TITLE       : {ctrl.get('control_title', 'N/A')}")
            report.append(f"   DOMAIN / SUBDOMAIN  : {ctrl.get('domain', 'N/A')} / {ctrl.get('subdomain', 'N/A')}")
            report.append(f"   SOURCE SHEET        : {ctrl.get('sheet', 'N/A')}")
            report.append(f"   COMPLIANCE STATUS   : {ctrl.get('compliance_status', 'UNKNOWN')}")

            if ctrl.get("compliance_status") == "COMPLIANT":
                report.append("   ASSESSMENT          : Control adequately meets applicable regulatory requirements.")
            else:
                report.append(f"   NON-COMPLIANCE REASON:")
                report.append(f"     - {ctrl.get('gaps', 'Insufficient control design or missing regulatory coverage.')}")

                report.append("   REQUIRED IMPROVEMENTS:")
                report.append(f"     - {ctrl.get('recommendation', 'Define and implement control enhancements aligned to regulation.')}")

            report.append(f"   REGULATORY MATCHES  : {ctrl.get('regulatory_matches', 0)}")

        # ------------------------------------------------------------------
        # DOMAIN THEMATIC OBSERVATIONS
        # ------------------------------------------------------------------
        report.append("\nDOMAIN-LEVEL OBSERVATIONS:")
        report.append("-" * 110)

        domain_gaps = [
            c for c in controls if c["compliance_status"] != "COMPLIANT"
        ]

        if not domain_gaps:
            report.append("✔ No significant gaps identified. Domain controls are operating effectively.")
        else:
            for gap in domain_gaps:
                report.append(
                    f"- Control {gap['control_reference']} requires enhancement to meet regulatory intent."
                )

        # ------------------------------------------------------------------
        # DOMAIN REMEDIATION ROADMAP
        # ------------------------------------------------------------------
        report.append("\nDOMAIN REMEDIATION RECOMMENDATIONS:")
        report.append("-" * 110)

        unique_recommendations = list({
            c["recommendation"]
            for c in controls
            if c.get("recommendation")
        })

        for rec in unique_recommendations:
            report.append(f"- {rec}")

    # ------------------------------------------------------------------
    # FINAL CONCLUSION
    # ------------------------------------------------------------------
    report.append("\n" + "=" * 110)
    report.append("OVERALL CONCLUSION")
    report.append("=" * 110)

    report.append(
        "This assessment evaluated the organization’s Risk Control Matrix against "
        "applicable technology risk regulations. While certain domains demonstrate "
        "adequate control maturity, multiple areas require enhancement to fully align "
        "with regulatory expectations. Addressing identified gaps will significantly "
        "reduce operational, compliance, and regulatory risk."
    )

    report.append("\n" + "=" * 110)
    report.append("END OF REPORT")
    report.append("=" * 110)

    return "\n".join(report)


def save_compliance_artifacts(
    results: Dict[str, Any],
    output_dir: str
) -> Dict[str, str]:
    """
    Save compliance analysis artifacts to disk.
    
    Args:
        results: Analysis results
        output_dir: Directory to save artifacts
    
    Returns:
        Dict with paths to saved files
    """
    try:
        os.makedirs(output_dir, exist_ok=True)
        
        artifacts = {}
        
        # Save report
        report_path = os.path.join(output_dir, "rcm_compliance_report.txt")
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(results.get('final_report', ''))
        artifacts['report'] = report_path
        logger.info(f"Saved report to: {report_path}")
        
        # Save JSON results
        import json
        json_path = os.path.join(output_dir, "rcm_compliance_results.json")
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)
        artifacts['json'] = json_path
        logger.info(f"Saved JSON to: {json_path}")
        
        # Save CSV of detailed results
        csv_path = os.path.join(output_dir, "rcm_compliance_details.csv")
        if results.get('success') and results.get('compliance_analysis'):
            detailed = results['compliance_analysis'].get('detailed_results', [])
            if detailed:
                import csv
                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=detailed[0].keys())
                    writer.writeheader()
                    writer.writerows(detailed)
                artifacts['csv'] = csv_path
                logger.info(f"Saved CSV to: {csv_path}")
        
        logger.info(f"Saved all compliance artifacts to: {output_dir}")
        
        return artifacts
        
    except Exception as e:
        logger.error(f"Failed to save artifacts: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return {}