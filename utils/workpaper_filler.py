"""
Streamlined Workpaper Filler
- Fills ICMP sheet with ALL controls
- Creates new Findings sheet with tabular summary
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


def safe_write_cell(ws, cell_address: str, value: Any):
    """Safely write to a cell, handling merged cells."""
    try:
        cell = ws[cell_address]
        for merged_range in list(ws.merged_cells.ranges):
            if cell.coordinate in merged_range:
                ws.unmerge_cells(str(merged_range))
                logger.debug(f"Unmerged cell range: {merged_range}")
                break
        ws[cell_address] = value
    except Exception as e:
        logger.error(f"Failed to write to cell {cell_address}: {e}")


def fill_workpaper_template(
    template_path: str,
    output_path: str,
    session_data: Dict[str, Any],
    analysis_results: List[Dict[str, Any]]
) -> str:
    """
    Fill workpaper with:
    1. ICMP sheet - Comprehensive details for ALL controls
    2. Findings sheet - Tabular summary of all findings
    
    Args:
        template_path: Path to Consolidated_WP_Template.xlsx
        output_path: Where to save filled workpaper
        session_data: Session info (test script, evidence files, etc.)
        analysis_results: LLM analysis for each control
    
    Returns:
        Path to filled workpaper
    """
    logger.info(f"Filling workpaper with {len(analysis_results)} controls")
    
    try:
        # Load template
        wb = openpyxl.load_workbook(template_path)
        
        # Get controls and results
        controls = session_data.get("controls", [])
        result_map = {r.get("control_id"): r for r in analysis_results}
        
        logger.info(f"Processing {len(controls)} controls")
        
        # Fill ICMP sheet with all controls
        fill_icmp_sheet_all_controls(wb, controls, result_map, session_data)
        
        # Create new Findings sheet with tabular summary
        create_findings_sheet(wb, controls, result_map, session_data)
        
        # Save
        wb.save(output_path)
        logger.info(f"Saved workpaper to: {output_path}")
        
        return output_path
        
    except Exception as e:
        logger.error(f"Failed to fill workpaper: {e}")
        raise


def fill_icmp_sheet_all_controls(
    wb,
    controls: List[Dict],
    result_map: Dict[str, Dict],
    session_data: Dict
):
    """Fill ICMP sheet with comprehensive details for ALL controls."""
    if "ICMP" not in wb.sheetnames:
        logger.warning("ICMP sheet not found, creating new one")
        wb.create_sheet("ICMP")
    
    ws = wb["ICMP"]
    
    # Clear existing content (keep first 5 rows for header)
    # First unmerge all cells to avoid MergedCell issues
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    
    # Now clear content
    for row in ws.iter_rows(min_row=6):
        for cell in row:
            try:
                cell.value = None
            except AttributeError:
                # Skip if cell is problematic
                pass
    
    # Header
    safe_write_cell(ws, "B1", "ICMP Controls - Comprehensive Testing Results")
    ws["B1"].font = Font(bold=True, size=14)
    
    safe_write_cell(ws, "B2", f"Testing Date: {datetime.now().strftime('%Y-%m-%d')}")
    safe_write_cell(ws, "B3", f"Total Controls Tested: {len(controls)}")
    
    # Summary statistics
    passed = sum(1 for r in result_map.values() if r.get("result") == "PASS")
    failed = sum(1 for r in result_map.values() if r.get("result") == "FAIL")
    partial = sum(1 for r in result_map.values() if r.get("result") == "PARTIAL")
    
    safe_write_cell(ws, "B4", f"Results: {passed} PASS | {failed} FAIL | {partial} PARTIAL | Pass Rate: {(passed/len(controls)*100):.1f}%" if len(controls) > 0 else "No controls tested")
    ws["B4"].font = Font(bold=True)
    
    # Detailed controls section - starting at row 7
    current_row = 7
    
    for idx, control in enumerate(controls, 1):
        control_id = control.get("control_id", f"CTRL-{idx}")
        result = result_map.get(control_id, {})
        
        # Section header with background
        header_cell = ws.cell(row=current_row, column=2, value=f"CONTROL #{idx}: {control_id}")
        header_cell.font = Font(bold=True, size=12, color="FFFFFF")
        header_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Merge across columns for header
        ws.merge_cells(f"B{current_row}:I{current_row}")
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Control Information Section
        info_section = ws.cell(row=current_row, column=2, value="CONTROL INFORMATION")
        info_section.font = Font(bold=True, size=11)
        info_section.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(f"B{current_row}:I{current_row}")
        current_row += 1
        
        # Control details table
        info_rows = [
            ("Control ID:", control_id),
            ("Control Name:", control.get("control_name", "N/A")),
            ("Control Type:", control.get("control_type", "Preventive")),
            ("Frequency:", control.get("frequency", "Ongoing")),
            ("Control Owner:", control.get("control_owner", "N/A")),
            ("Description:", control.get("control_description", "N/A")[:300]),
        ]
        
        for label, value in info_rows:
            label_cell = ws.cell(row=current_row, column=2, value=label)
            label_cell.font = Font(bold=True)
            label_cell.alignment = Alignment(horizontal="right", vertical="top")
            
            value_cell = ws.cell(row=current_row, column=3, value=str(value))
            value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.merge_cells(f"C{current_row}:I{current_row}")
            
            if label == "Description:":
                ws.row_dimensions[current_row].height = 40
            else:
                ws.row_dimensions[current_row].height = 20
                
            current_row += 1
        
        # Test Procedures Section
        current_row += 1
        proc_section = ws.cell(row=current_row, column=2, value="TEST PROCEDURES")
        proc_section.font = Font(bold=True, size=11)
        proc_section.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        ws.merge_cells(f"B{current_row}:I{current_row}")
        current_row += 1
        
        test_steps = control.get("test_steps", "No test steps defined").split("\n")
        for step_idx, step in enumerate(test_steps[:10], 1):  # Max 10 steps
            if step.strip():
                step_cell = ws.cell(row=current_row, column=2, value=f"{step_idx}.")
                step_cell.font = Font(bold=True)
                
                step_text = ws.cell(row=current_row, column=3, value=step.strip()[:200])
                step_text.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.merge_cells(f"C{current_row}:I{current_row}")
                ws.row_dimensions[current_row].height = 30
                current_row += 1
        
        # Sample & Evidence Section
        current_row += 1
        ws.cell(row=current_row, column=2, value="Sample Size:").font = Font(bold=True)
        ws.cell(row=current_row, column=3, value=control.get("sample_size", "N/A"))
        current_row += 1
        
        ws.cell(row=current_row, column=2, value="Evidence Files:").font = Font(bold=True)
        uploaded_files = session_data.get("uploaded_files", {})
        if uploaded_files:
            evidence_list = ", ".join(list(uploaded_files.keys())[:5])
            evidence_cell = ws.cell(row=current_row, column=3, value=evidence_list)
            evidence_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"C{current_row}:I{current_row}")
            ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Test Results Section
        current_row += 1
        results_section = ws.cell(row=current_row, column=2, value="TEST RESULTS")
        results_section.font = Font(bold=True, size=11, color="FFFFFF")
        results_section.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        ws.merge_cells(f"B{current_row}:I{current_row}")
        current_row += 1
        
        # Result with color coding
        ws.cell(row=current_row, column=2, value="Result:").font = Font(bold=True)
        result_val = result.get("result", "NOT_TESTED")
        result_cell = ws.cell(row=current_row, column=3, value=result_val)
        result_cell.font = Font(bold=True, size=12)
        
        if result_val == "PASS":
            result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            result_cell.font = Font(bold=True, size=12, color="006100")
        elif result_val == "FAIL":
            result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            result_cell.font = Font(bold=True, size=12, color="9C0006")
        elif result_val == "PARTIAL":
            result_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            result_cell.font = Font(bold=True, size=12, color="9C6500")
        
        ws.row_dimensions[current_row].height = 25
        current_row += 1
        
        # Observation
        ws.cell(row=current_row, column=2, value="Observation:").font = Font(bold=True)
        obs_cell = ws.cell(row=current_row, column=3, value=result.get("observation", "No observation recorded"))
        obs_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"C{current_row}:I{current_row}")
        ws.row_dimensions[current_row].height = 35
        current_row += 1
        
        # Exceptions
        exceptions = result.get("exceptions", [])
        if exceptions:
            ws.cell(row=current_row, column=2, value="Exceptions Found:").font = Font(bold=True, color="C00000")
            ws.cell(row=current_row, column=3, value=f"{len(exceptions)} exception(s)")
            ws["C" + str(current_row)].font = Font(bold=True, color="C00000")
            current_row += 1
            
            for exc_idx, exc in enumerate(exceptions[:10], 1):  # Max 10 exceptions
                exc_num = ws.cell(row=current_row, column=2, value=f"{exc_idx}.")
                exc_num.alignment = Alignment(horizontal="right")
                
                exc_cell = ws.cell(row=current_row, column=3, value=str(exc)[:250])
                exc_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.merge_cells(f"C{current_row}:I{current_row}")
                ws.row_dimensions[current_row].height = 30
                current_row += 1
        else:
            ws.cell(row=current_row, column=2, value="Exceptions:").font = Font(bold=True)
            ws.cell(row=current_row, column=3, value="None")
            current_row += 1
        
        # Recommendation
        current_row += 1
        ws.cell(row=current_row, column=2, value="Recommendation:").font = Font(bold=True)
        rec_cell = ws.cell(row=current_row, column=3, value=result.get("recommendation", "No recommendation"))
        rec_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.merge_cells(f"C{current_row}:I{current_row}")
        ws.row_dimensions[current_row].height = 30
        current_row += 1
        
        # Impact
        impact = result.get("impact", "MEDIUM")
        ws.cell(row=current_row, column=2, value="Impact:").font = Font(bold=True)
        impact_cell = ws.cell(row=current_row, column=3, value=impact)
        impact_cell.font = Font(bold=True)
        
        if impact == "HIGH":
            impact_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            impact_cell.font = Font(bold=True, color="9C0006")
        elif impact == "MEDIUM":
            impact_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            impact_cell.font = Font(bold=True, color="9C6500")
        else:  # LOW
            impact_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            impact_cell.font = Font(bold=True, color="006100")
        
        current_row += 1
        
        # Separator
        current_row += 2
    
    # Set column widths
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 70
    for col in ['D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 12
    
    logger.info(f"Filled ICMP sheet with {len(controls)} controls")


def create_findings_sheet(
    wb,
    controls: List[Dict],
    result_map: Dict[str, Dict],
    session_data: Dict
):
    """Create new Findings sheet with tabular summary of all findings."""
    
    # Create new sheet
    if "Findings" in wb.sheetnames:
        del wb["Findings"]
    
    ws = wb.create_sheet("Findings", 0)  # Insert as first sheet
    
    # Header
    ws["A1"] = "AUDIT FINDINGS SUMMARY"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:L1")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws["A2"] = f"Testing Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws.merge_cells("A2:L2")
    
    ws["A3"] = f"Total Controls: {len(controls)}"
    ws.merge_cells("A3:L3")
    
    # Statistics
    passed = sum(1 for r in result_map.values() if r.get("result") == "PASS")
    failed = sum(1 for r in result_map.values() if r.get("result") == "FAIL")
    partial = sum(1 for r in result_map.values() if r.get("result") == "PARTIAL")
    
    ws["A4"] = f"Results: {passed} PASS | {failed} FAIL | {partial} PARTIAL | Pass Rate: {(passed/len(controls)*100):.1f}%"
    ws["A4"].font = Font(bold=True)
    ws.merge_cells("A4:L4")
    
    # Table headers
    headers = [
        "#",
        "Control ID", 
        "Control Name",
        "Type",
        "Frequency",
        "Result",
        "Sample Size",
        "Exceptions",
        "Impact",
        "Observation",
        "Recommendation",
        "Evidence Files"
    ]
    
    row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Freeze header row
    ws.freeze_panes = "A7"
    
    # Fill data rows
    row = 7
    for idx, control in enumerate(controls, 1):
        control_id = control.get("control_id", f"CTRL-{idx}")
        result = result_map.get(control_id, {})
        
        # Row data
        data = [
            idx,  # #
            control_id,  # Control ID
            control.get("control_name", "")[:60],  # Control Name
            control.get("control_type", "Preventive"),  # Type
            control.get("frequency", "Ongoing"),  # Frequency
            result.get("result", "NOT_TESTED"),  # Result
            control.get("sample_size", "N/A"),  # Sample Size
            len(result.get("exceptions", [])),  # Exceptions count
            result.get("impact", "MEDIUM"),  # Impact
            result.get("observation", "")[:150],  # Observation
            result.get("recommendation", "")[:150],  # Recommendation
            ", ".join(list(session_data.get("uploaded_files", {}).keys())[:3])  # Evidence
        ]
        
        for col_idx, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code Result column (column 6)
            if col_idx == 6:
                if value == "PASS":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, color="006100")
                elif value == "FAIL":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True, color="9C0006")
                elif value == "PARTIAL":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(bold=True, color="9C6500")
            
            # Color code Impact column (column 9)
            if col_idx == 9:
                if value == "HIGH":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(bold=True, color="9C0006")
                elif value == "MEDIUM":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(bold=True, color="9C6500")
                elif value == "LOW":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(bold=True, color="006100")
        
        # Set row height
        ws.row_dimensions[row].height = 40
        row += 1
    
    # Set column widths
    column_widths = {
        'A': 6,   # #
        'B': 15,  # Control ID
        'C': 30,  # Control Name
        'D': 12,  # Type
        'E': 12,  # Frequency
        'F': 12,  # Result
        'G': 12,  # Sample Size
        'H': 10,  # Exceptions
        'I': 10,  # Impact
        'J': 35,  # Observation
        'K': 35,  # Recommendation
        'L': 25,  # Evidence
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    logger.info(f"Created Findings sheet with {len(controls)} findings")


# For backward compatibility
fill_workpaper_template_all_controls = fill_workpaper_template