"""
Test Script Parser
Parses Excel test scripts to extract control testing requirements
NO API CODE - pure parsing logic
"""

import openpyxl
from typing import List, Dict, Optional, Any
from pathlib import Path
import logging

logger = logging.getLogger(__name__)

# Column mapping - flexible matching
COLUMN_MAP = {
    "control_id": ["control id", "control ref", "control #", "id", "reference"],
    "risk_statement": ["risk statement", "risk", "risk desc", "control risk statement"],
    "control_description": ["control description", "control activity", "what control does", "description"],
    "control_owner": ["control owner", "owner", "accountability", "performed by", "reviewer"],
    "frequency": ["frequency", "adjusted frequency", "control frequency"],
    "test_objective": ["test objective", "objective", "what auditor", "objective addressed"],
    "test_steps": ["test step", "procedure", "step-by-step", "test procedures", "procedures performed"],
    "evidence_required": ["evidence required", "evidence", "what to collect", "evidence needed"],
    "sample_size": ["sample size", "how much", "population", "sample"],
}


def parse_test_script(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse Excel test script to extract control testing requirements.
    
    Args:
        file_path: Path to test script Excel file
    
    Returns:
        List of control dicts with standardized fields
    
    Raises:
        ValueError: If file cannot be parsed or has invalid structure
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        logger.info(f"Loaded workbook: {file_path}, sheets: {wb.sheetnames}")
        
        # Try first sheet by default
        ws = wb.active
        logger.info(f"Using sheet: {ws.title}")
        
        # Find header row
        header_row_idx = find_header_row(ws)
        if not header_row_idx:
            raise ValueError("Could not find header row in test script")
        
        logger.info(f"Found header row at index: {header_row_idx}")
        
        # Build column mapping
        column_mapping = build_column_mapping(ws, header_row_idx)
        logger.info(f"Column mapping: {column_mapping}")
        
        if not column_mapping.get("control_id") and not column_mapping.get("control_description"):
            raise ValueError("Could not find control_id or control_description columns")
        
        # Extract rows
        controls = extract_controls(ws, header_row_idx, column_mapping)
        
        logger.info(f"Parsed {len(controls)} controls from test script")
        
        return controls
        
    except Exception as e:
        logger.error(f"Failed to parse test script: {e}")
        raise ValueError(f"Test script parsing failed: {e}")


def find_header_row(ws) -> Optional[int]:
    """
    Find the row containing column headers.
    Looks for rows with multiple text cells containing expected keywords.
    """
    # Search first 20 rows
    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_values = [cell.value for cell in ws[row_idx]]
        
        # Filter None and convert to lowercase strings
        text_cells = [
            str(v).lower() for v in row_values 
            if v is not None and str(v).strip()
        ]
        
        # Must have at least 3 text cells
        if len(text_cells) < 3:
            continue
        
        # Check if any expected column names are present
        header_keywords = ["control", "risk", "test", "objective", "evidence", "procedure"]
        matches = sum(
            1 for cell in text_cells 
            for keyword in header_keywords 
            if keyword in cell
        )
        
        if matches >= 2:  # At least 2 header keywords found
            return row_idx
    
    return None


def build_column_mapping(ws, header_row_idx: int) -> Dict[str, int]:
    """
    Build mapping from standardized field names to column indices.
    
    Returns:
        Dict mapping field_name -> column_index (0-based)
    """
    header_row = ws[header_row_idx]
    column_mapping = {}
    
    # Get header values
    headers = [
        str(cell.value).lower().strip() if cell.value else "" 
        for cell in header_row
    ]
    
    # Match headers to standard fields
    for field_name, match_terms in COLUMN_MAP.items():
        for col_idx, header_text in enumerate(headers):
            if not header_text:
                continue
            
            # Check if any match term is in header
            for match_term in match_terms:
                if match_term in header_text:
                    column_mapping[field_name] = col_idx
                    break
            
            if field_name in column_mapping:
                break
    
    return column_mapping


def extract_controls(
    ws, 
    header_row_idx: int, 
    column_mapping: Dict[str, int]
) -> List[Dict[str, Any]]:
    """
    Extract control rows from worksheet.
    
    Args:
        ws: Worksheet
        header_row_idx: Row index of headers
        column_mapping: Field to column index mapping
    
    Returns:
        List of control dicts
    """
    controls = []
    
    # Process rows after header
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row = ws[row_idx]
        
        # Extract values based on column mapping
        control = {}
        has_content = False
        
        for field_name, col_idx in column_mapping.items():
            if col_idx < len(row):
                cell_value = row[col_idx].value
                
                # Clean and convert
                if cell_value is not None:
                    if isinstance(cell_value, str):
                        cell_value = cell_value.strip()
                    if cell_value:
                        control[field_name] = str(cell_value)
                        has_content = True
        
        # Only include rows with content
        if has_content and (control.get("control_id") or control.get("control_description")):
            # Generate control_id if missing
            if not control.get("control_id"):
                control["control_id"] = f"CTL-{len(controls) + 1:03d}"
            
            # Set defaults for missing fields
            control.setdefault("risk_statement", "Not specified")
            control.setdefault("control_description", "Not specified")
            control.setdefault("control_owner", "Not specified")
            control.setdefault("frequency", "Not specified")
            control.setdefault("test_objective", "Verify control effectiveness")
            control.setdefault("test_steps", "Review evidence and assess compliance")
            control.setdefault("evidence_required", "Supporting documentation")
            control.setdefault("sample_size", "As applicable")
            
            # Initialize result fields (filled later by analyzer)
            control["result"] = None
            control["observation"] = None
            control["kb1_reference"] = None
            control["kb2_reference"] = None
            control["recommendation"] = None
            control["exceptions"] = []
            
            controls.append(control)
    
    return controls


def validate_controls(controls: List[Dict[str, Any]]) -> List[str]:
    """
    Validate parsed controls for completeness.
    
    Returns:
        List of validation warnings
    """
    warnings = []
    
    if not controls:
        warnings.append("No controls found in test script")
        return warnings
    
    for idx, control in enumerate(controls, 1):
        control_id = control.get("control_id", f"Row {idx}")
        
        if not control.get("control_description") or control["control_description"] == "Not specified":
            warnings.append(f"{control_id}: Missing control description")
        
        if not control.get("evidence_required") or control["evidence_required"] == "Supporting documentation":
            warnings.append(f"{control_id}: Evidence requirement is vague")
    
    return warnings


def get_test_script_summary(controls: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Get summary statistics from parsed controls.
    
    Returns:
        Summary dict with counts and breakdowns
    """
    if not controls:
        return {
            "total_controls": 0,
            "frequency_breakdown": {},
            "owners": []
        }
    
    frequency_counts = {}
    owners = set()
    
    for control in controls:
        freq = control.get("frequency", "Unknown")
        frequency_counts[freq] = frequency_counts.get(freq, 0) + 1
        
        owner = control.get("control_owner", "Unknown")
        if owner != "Not specified":
            owners.add(owner)
    
    return {
        "total_controls": len(controls),
        "frequency_breakdown": frequency_counts,
        "owners": sorted(list(owners)),
        "controls_with_evidence_spec": sum(
            1 for c in controls 
            if c.get("evidence_required") and c["evidence_required"] != "Supporting documentation"
        )
    }
